# -*- coding: utf-8 -*-
"""
Created on Mon Dec  2 11:46:53 2019

In this project, you’ll write a program to update cells in a spreadsheet of
produce sales. Your program will look through the spreadsheet, find specific
kinds of produce, and update their prices.
Each row represents an individual sale. The columns are the type of
produce sold (A), the cost per pound of that produce (B), the number of
pounds sold (C), and the total revenue from the sale (D). The TOTAL column
is set to the Excel formula =ROUND(B3*C3, 2), which multiplies the
cost per pound by the number of pounds sold and rounds the result to the
nearest cent. With this formula, the cells in the TOTAL column will automatically
update themselves if there is a change in column B or C.

@author: HP PC
"""

import openpyxl, os

os.chdir('C:\\Users\\HP PC\\Downloads')
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

PRICE_UPDATES = {'Garlic': 3.07,
				 'Celery': 1.19,
				 'Lemon': 1.27}
# TODO: Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row):
	produceName = sheet.cell(row=rowNum, column=1).value
	if produceName in PRICE_UPDATES:
		sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
		
wb.save('updatedProduceSales.xlsx')
